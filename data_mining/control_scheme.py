import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook(r'C:\Users\lln\Desktop\保定数据研究\8点-8点05\信号控制方案\研究交叉口信号控制方案.xlsx')
wa = openpyxl.load_workbook(r'C:\Users\lln\Desktop\保定数据研究\8点-8点05\信号控制方案\研究交叉口信号控制方案_v2.xlsx')
sheet = wb.active
sheet1 = wa.active
wd = Workbook()
ws = wd.active

list1=[]
list2=[]
list3=[]
list4=[]
list5=[]
dict1={}
aa=0
j=1

def judge(a):
    global x
    if a=="1-5"or a=="0-6" or a=="0-5":
        x=1
    elif a=="0,6"or a=='6':
        x=0
    return x

for i in range(2,sheet1.max_row+1):
    date=sheet1['E'+str(i)].value[7:]
    if judge(date)==1:
        list1.append(i)
print(list1)
print(len(list1))
for a in list1:
    point=sheet1['I'+str(a)].value
    list2.append(point)
    time=eval(sheet1['F'+str(a)].value)
    for fea in range(len(time)):
        ss=time[fea]['duration']
        aa=ss+aa
        if aa>480:
            scheme=time[fea]['scheme_id']
            list3.append(scheme)
            aa=0
            break
print(list2)#交叉口id
print(len(list2))
print(list3)#scheme_id
print(len(list3))
for ab in range(len(list2)):
    dict1[list2[ab]]=list3[ab]
print(dict1)#各交叉口id对应scheme

for abc in range(2,sheet.max_row+1):
    arcgis=sheet['J'+str(abc)].value
    scheme=int(sheet['C'+str(abc)].value)
    for abb in range(len(list2)):
        if arcgis==list2[abb] and scheme==list3[abb]:
            list4.append(abc)
print(list4)#提取对应scheme行数
print(len(list4))
for hang in list4:
    signal=eval(sheet["F"+str(hang)].value)
    ws["A1"]="No"+str(sheet["J"+str(hang)].value)
    ws["A2"]="Cycle="+sheet["D"+str(hang)].value+"s"
    ws["B2"]="Offset="+sheet["E"+str(hang)].value+"s"
    ws["A3"]="order"
    ws["B3"] = "green_time"
    ws["C3"] = "yellow_time"
    ws["D3"] = "red_time"
    ws["E3"] = None
    ws["F3"] = "barrier"
    ws["G3"] = "ring"
    ws["H3"] = "phase_id"
    for aa in range(len(signal)):
        ws["A"+str(j+3)]=signal[aa]['order']
        ws["B" + str(j + 3)] = signal[aa]['green_time']
        ws["C" + str(j + 3)] = signal[aa]['yellow_time']
        ws["D" + str(j + 3)] = signal[aa]['red_time']
        ws["E" + str(j + 3)] = signal[aa]['green_time']+signal[aa]['yellow_time']+signal[aa]['red_time']
        ws["F" + str(j + 3)] = signal[aa]['barrier']
        ws["G" + str(j + 3)] = signal[aa]['ring']
        ws["H" + str(j + 3)] = signal[aa]['phase_id']
        j=j+1
    wd.save("No."+str(sheet["J"+str(hang)].value)+'.xlsx')
    j=1
    wd = Workbook()
    ws = wd.active