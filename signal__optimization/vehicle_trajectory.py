import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook(r'C:\Users\lln\Desktop\保定数据研究\9点30-10点30\研究交叉口提取及arcgis,time(9点30-10点30).xlsx')
sheet = wb.active
wd = Workbook()
ws = wd.active

result=[[]]
pointres=[]
timeres=[]
lst=[]
ln=[]
pointlist=[]
timelist=[]
c=0

for i in range(2,sheet.max_row+1):
    tra=sheet['D'+str(i)].value
    point=sheet['K'+str(i)].value
    time=sheet['L'+str(i)].value
    lst.append(tra)
    pointlist.append(point)
    timelist.append(time)
print(lst)#所有轨迹点

for i in range(len(lst)-1):
    if lst[i]+1==lst[i+1]:
        result[-1].append(lst[i])
    else:
        result[-1].append(lst[i])
        result.append([])
        result[-1].append(lst[i])
b=[result[0]]
for s in range(1,len(result)):
    a=result[s][1:]
    b.append(a)
if b[-1][-1]+1==lst[-1]:
    b[-1].append(lst[-1])
print(b)#相邻轨迹点分段

ln.append(0)
for i in range(len(b)):
    c=c+len(b[i])
    ln.append(c)
ln=ln[:-1]
print(ln)#分段点

for i in range(len(ln)-1):
    pointres.append(pointlist[ln[i]:ln[i+1]])
pointres.append(pointlist[ln[i+1]:])
print(pointres)#卡口分段

for i in range(len(ln)-1):
    timeres.append(timelist[ln[i]:ln[i+1]])
timeres.append(timelist[ln[i+1]:])
print(timeres)#时间分段

for i in range(len(pointres)):
    if len(pointres[i])==1:
        pointres[i]=[]
        timeres[i]=[]#去除单点路径

for i in range(len(pointres)):
    ws['A'+str(i+1)]=str(pointres[i])
    ws['B'+str(i+1)]=str(timeres[i])
wd.save("车辆路径及时间.xlsx")