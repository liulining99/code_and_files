import json
import optparse
import sys
import time
import traci
import os
from sumolib import checkBinary
from openpyxl import Workbook

j=0
allnum=0
s=0
edgelengthlist=[]
edgevehicle=[]
networknum=[]
edgeflow=[]
flow=[]
thetime=[]
lastvehIDs=[]
aveflow=[]
alllength=0
wd = Workbook()
ws = wd.active

for i in range(470):
    lastvehIDs.append([])
    edgeflow.append([])
print(lastvehIDs)
if 'SUMO_HOME' in os.environ:
    tools=os.path.join(os.environ['SUMO_HOME'],'tools')
    sys.path.append(tools)
else:
    sys.exit("please declare environment variable 'SUMO_HOME'")

if_show_gui = True

if not if_show_gui:
    sumoBinary=checkBinary('sumo')
else:
    sumoBinary=checkBinary('sumo-gui')

sumocfgfile=r"C:\Users\lln\Desktop\zhongshan-univer project\junction-vehicle.sumocfg"
traci.start([sumoBinary,"-c",sumocfgfile])
#读取各路段长度
edgeID=traci.edge.getIDList()
print(traci.edge.getIDList())
print(len(edgeID))
edgeID=list(edgeID[:470])
print(edgeID)
for edge in edgeID:
    lane=edge+'_0'
    lanelength=traci.lane.getLength(lane)
    edgelengthlist.append(lanelength)
print(edgelengthlist)
print(len(edgeID))
print(len(edgelengthlist))

for leng in edgelengthlist:
    alllength=alllength+leng

for step in range(0,10000):
    traci.simulationStep(step+1)
    simulation_current_time=traci.simulation.getTime()
    # print(simulation_current_time)
    thetime.append(simulation_current_time)
    for x in edgeID:
        edgevehicle.append(list(traci.edge.getLastStepVehicleIDs(x)))
    for i in range(len(edgeID)):
        for a in lastvehIDs[i]:
            if a not in edgevehicle[i]:
                edgeflow[i].append(a)
        flow.append(len(edgeflow[i]))
        edgeflow[i]=[]
    # print(flow)
    # print(edgevehicle)
    for q in edgevehicle:
        allnum = allnum + len(q)  # 路网内累计车辆数
    for aa in range(len(edgeID)):
        s=s+flow[aa]*edgelengthlist[aa]/alllength
    aveflow.append(s)
    networknum.append(allnum)
    lastvehIDs=edgevehicle
    edgevehicle=[]
    flow=[]
    allnum=0
    s=0
for i in range(len(networknum)):
    ws['A'+str(i+1)]=thetime[i]
    ws['B'+str(i+1)]=networknum[i]
    ws['C'+str(i+1)]=aveflow[i]
wd.save('数据2.xlsx')
traci.close()