from openpyxl import Workbook
import openpyxl
import networkx as nx
from matplotlib import pyplot as plt

wd=openpyxl.load_workbook(r'C:\Users\lln\Desktop\baoding_sumo_nofail_v\路段平均速度.xlsx')
sheet1=wd.active
wq=openpyxl.load_workbook(r'C:\Users\lln\Desktop\baoding_result\arcgis与sumo对应关系.xlsx')
sheet2=wq.active
ws=openpyxl.load_workbook(r'C:\Users\lln\Desktop\保定研究路网关系界定\线0-1矩阵-改.xlsx')
sheet3=ws.active
wn = Workbook()
wb = wn.active

t=1
road=[]
road_arcgis=[]
road_speed=[]
valid_road=[]
valid_road_speed=[]
edgelist=[]
edge_all=[]
dict1={}
tup1=()
edgelist1 = []
numvalidline=[]
maxGsize=[]
numG=[]
#创建路段对应字典
for i in range(2,sheet2.max_row+1):
    key=str(sheet2["A"+str(i)].value)
    value=int(sheet2["B"+str(i)].value)
    dict1[key] = value
print(dict1)
#创建所有路段的邻接关系元组
for i in range(1,sheet3.max_row+1):
    for j in range(1,sheet3.max_row+1):
        x = sheet3.cell(row=i, column=j).value
        if x==1:
            tup1=(i-1,j-1)
            tup2=(i+299,j+299)
            tup3=(i-1,i+299)
            tup4=(i+299,i-1)
            edge_all.append(tup1)
            edge_all.append(tup2)
            edge_all.append(tup3)
            edge_all.append(tup4)
print(edge_all)

#road、road_speed添加所有路段编号、速度
for i in range(2,sheet1.max_row+1):
    speed=float(sheet1["B"+str(i)].value)*3.6
    line=sheet1["A"+str(i)].value
    road.append((line))
    road_speed.append(speed)
print(road)
print(road_speed)
#输出road列表对应的arcgis编号列表
for abcd in road:
    bbb=dict1.get(abcd,None)
    road_arcgis.append(int(bbb))
print(road_arcgis)
road_arcgis1=list(filter(None,road_arcgis))
print(road_arcgis1)
# valid_road、valid_road_speed添加对应于某一速度阈值下的有效路段编号、速度
for a in range(80):#速度最大值
    for s in range(len(road_arcgis1)):#路段数
        if road_speed[s]>a:#判定是否有效
            valid_road_speed.append(road_speed[s])
            aaa=dict1.get(str(road[s]),None)
            valid_road.append(aaa)
    valid_road_node=list(filter(None,valid_road))
    # print(valid_road_node)
    numvalidline.append(len(valid_road_node))
    # print(len(valid_road_node))#输出有效路段数
    G = nx.Graph()
    for i in valid_road_node:
        for aa in edge_all:
            if abs(i) in aa:
                edgelist.append(aa)
    # print(edgelist)
    edgelist1 = []
    for ab in edgelist:
        if ab[0] in valid_road_node and ab[1] in valid_road_node:
            edgelist1.append(ab)
    G.add_nodes_from(valid_road_node)
    G.add_edges_from(edgelist1)
    # print(valid_road_node)
    # print(edgelist1)
    # print(max(nx.connected_components(G)))
    maxGsize.append(max(nx.connected_components(G)))  # 最大集群包含节点
    valid_road.clear()
    valid_road_speed.clear()
    edgelist.clear()
    edgelist1.clear()

for a in range(80):
    for s in range(len(road_arcgis1)):
        if road_speed[s]>a:
            valid_road_speed.append(road_speed[s])
            aaa=dict1.get(str(road[s]),None)
            valid_road.append(aaa)
    valid_road_node=list(filter(None,valid_road))
    # print(len(valid_road_node))#输出有效路段数
    G = nx.Graph()
    for i in valid_road_node:
        for aa in edge_all:
            if abs(i) in aa:
                edgelist.append(aa)
    for ab in edgelist:
        if ab[0] in valid_road_node and ab[1] in valid_road_node:
            edgelist1.append(ab)
    G.add_nodes_from(road_arcgis1)
    G.add_edges_from(edgelist1)
    numG.append(nx.number_connected_components(G))  # 集群数
    valid_road.clear()
    valid_road_speed.clear()
    edgelist.clear()
    edgelist1.clear()

maxGsize1=[]
for abcde in maxGsize:
    maxGsize1.append(len(abcde))
print(numvalidline)
print(numG)
print(maxGsize1)
v=range(0,80)

plt.figure()
plt.plot(v,numvalidline)
plt.xticks([0,10,20,30,40,50,60,70,80,90,100])
plt.xlim(0,100)
plt.yticks([0,50,100,150,200,250])
plt.ylim(0,250)
plt.xlabel("vt")
plt.ylabel("num valid road")

plt.figure()
plt.plot(v,numG)
plt.xticks([0,10,20,30,40,50,60,70,80,90,100])
plt.xlim(0,100)
plt.yticks([0,50,100,150,200])
plt.ylim(0,250)
plt.xlabel("vt")
plt.ylabel("num cluster")

plt.figure()
plt.plot(v,maxGsize1)
plt.xticks([0,10,20,30,40,50,60,70,80,90,100])
plt.xlim(0,100)
plt.yticks([0,50,100,150,200,250])
plt.ylim(0,250)
plt.xlabel("vt")
plt.ylabel("size of maxcluster")

plt.show()

for i in range(len(maxGsize1)):
    wb['A' + str(t)] =v[i]
    wb['B' + str(t)] =numvalidline[i]
    wb['C' + str(t)] =numG[i]
    wb['D' + str(t)] =maxGsize1[i]
    t=t+1
wn.save('复杂网络信息.xlsx')