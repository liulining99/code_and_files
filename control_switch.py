import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook(r'C:\Users\lln\Desktop\保定数据研究\8点-8点05\信号控制方案\各交叉口控制方案\No.38.xlsx')
sheet = wb.active
wd = Workbook()
ws = wd.active

ring_list1 = []
ring_list2 = []
barrier_list = []
phase_id_list1 = []
phase_id_list2 = []
G_list1 = []
Y_list1 = []
R_list1 = []
G_list2 = []
Y_list2 = []
R_list2 = []
order_list1 = []
order_list2 = []
no0 = []
index_list1 = []
index_list2 = []
ss1 = []
cc1 = []
bb1 = []
cycle_list1 = []
cycle_list2 = []
cycle_list = []
dict1 = {}
dict2 = {}
dict3 = {}
s = 0
ss = 0
cc = 0
cycle = eval(sheet['A2'].value[6:-1])
print(cycle)
cycle1 = 0
cycle2 = 0
t0 = 0
origin = ["G", "r", "r", "r", "G", "r", "r", "r", "G", "r", "r", "r", "G", "r", "r", "r"]

for a in range(4, sheet.max_row + 1):
    if sheet["H" + str(a)].value != 0:
        no0.append(a)
x = no0[-1]
for i in range(4, x + 1):
    ring = sheet["G" + str(i)].value
    if ring == 0:
        index_list1.append(i)
    elif ring == 1:
        index_list2.append(i)
for b in range(index_list1[0], index_list1[-1] + 1):
    G_list1.append(sheet["B" + str(b)].value)
    Y_list1.append(sheet["C" + str(b)].value)
    R_list1.append(sheet["D" + str(b)].value)
    phase_id_list1.append(sheet["H" + str(b)].value)
    order_list1.append(sheet["A" + str(b)].value)

for ab in range(index_list2[0], index_list2[-1] + 1):
    G_list2.append(sheet["B" + str(ab)].value)
    Y_list2.append(sheet["C" + str(ab)].value)
    R_list2.append(sheet["D" + str(ab)].value)
    phase_id_list2.append(sheet["H" + str(ab)].value)
    order_list2.append(sheet["A" + str(ab)].value)
dict1["G"] = G_list1
dict1["Y"] = Y_list1
dict1["R"] = R_list1
dict1["phase"] = phase_id_list1
dict1["order"] = order_list1
dict2["G"] = G_list2
dict2["Y"] = Y_list2
dict2["R"] = R_list2
dict2["phase"] = phase_id_list2
dict2["order"] = order_list2
print(dict1, dict2)  # 记得要考虑order相等的情况

for i in range(len(order_list1)):
    ss = ss + dict1["G"][i]
    cc = cc + dict2["G"][i]
    ss1.append(ss)
    cc1.append(cc)
    ss = ss + dict1["Y"][i]
    cc = cc + dict2["Y"][i]
    ss1.append(ss)
    cc1.append(cc)
    ss = ss + dict1["R"][i]
    cc = cc + dict2["R"][i]
    ss1.append(ss)
    cc1.append(cc)
a1 = sorted(ss1 + cc1)
aa1 = list(set(a1))
aa1.sort(key=a1.index)
print(len(aa1))  # len(aa1)为灯色组合数量

aa1.insert(0, 0)
print(aa1)  # 周期计算

for i in range(len(aa1) - 1):
    bb1.append(aa1[i + 1] - aa1[i])
print(bb1)  # 各灯色时间

# ring1的每秒灯色显示
for a in range(len(order_list1)):
    for t in range(cycle):
        cycle1 = 0
        cycle1 = cycle1 + dict1['G'][a]
        aaaa = cycle1
        if t < cycle1:
            abb = str(dict1['phase'][a]) + "G"
        elif t == cycle1:
            abb = str(dict1['phase'][a]) + "Y"
        cycle1 = cycle1 + dict1['Y'][a]
        bbbb = cycle1
        if aaaa < t < cycle1:
            abb = str(dict1['phase'][a]) + "Y"
        elif t == cycle1:
            if dict1['R'][a] != 0:
                abb = str(dict1['phase'][a]) + "r"
            if dict1['R'][a] == 0:
                abb = str(dict1['phase'][a + 1]) + "G"
        cycle1 = cycle1 + dict1['R'][a]
        if bbbb < t < cycle1:
            abb = str(dict1['phase'][a]) + "r"
        elif t == cycle1 and a != len(order_list1) - 1:
            abb = str(dict1['phase'][a + 1]) + "G"
        cycle_list1.append(abb)
        # print(t,abb,aaaa,bbbb,cycle1)
        if t > cycle1 - 2:
            break
print(cycle_list1)
print(len(cycle_list1))

# ring2的每秒灯色显示
for a in range(len(order_list1)):
    for t in range(cycle):
        cycle2 = 0
        cycle2 = cycle2 + dict2['G'][a]
        aaaa = cycle2
        if t < cycle2:
            abb = str(dict2['phase'][a]) + "G"
        elif t == cycle2:
            abb = str(dict2['phase'][a]) + "Y"
        cycle2 = cycle2 + dict2['Y'][a]
        bbbb = cycle2
        if aaaa < t < cycle2:
            abb = str(dict2['phase'][a]) + "Y"
        elif t == cycle2:
            if dict2['R'][a] != 0:
                abb = str(dict2['phase'][a]) + "r"
            if dict2['R'][a] == 0:
                abb = str(dict2['phase'][a + 1]) + "G"
        cycle2 = cycle2 + dict2['R'][a]
        if bbbb < t < cycle2:
            abb = str(dict2['phase'][a]) + "r"
        elif t == cycle2 and a != len(order_list1) - 1:
            abb = str(dict2['phase'][a + 1]) + "G"
        cycle_list2.append(abb)
        # print(t,abb,aaaa,bbbb,cycle2)
        if t > cycle2 - 2:
            break
print(cycle_list2)
print(len(cycle_list2))

# ring1与ring2合并
for i in range(len(cycle_list1)):
    sss = cycle_list1[i] + cycle_list2[i]
    cycle_list.append(sss)
print(cycle_list)
cyclelist = list({}.fromkeys(cycle_list).keys())
print(cyclelist)

dict3[1] = [10, 11]
dict3[2] = [1]
dict3[3] = [14, 15]
dict3[4] = [5]
dict3[5] = [2, 3]
dict3[6] = [9]
dict3[7] = [6, 7]
dict3[8] = [13]
print(dict3)  # 对应字符位置

scheme = []
for i in range(len(cyclelist)):
    for b in dict3[int(cyclelist[i][0])]:
        origin[b] = cyclelist[i][1]
    for c in dict3[int(cyclelist[i][2])]:
        origin[c] = cyclelist[i][3]
    varied = ''.join(origin)
    scheme.append(varied)
    origin = ["G", "r", "r", "r", "G", "r", "r", "r", "G", "r", "r", "r", "G", "r", "r",
              "r"]  # 输出sumo控制方案(scheme),bb1是时长
print(bb1)
print(scheme)
dashu = []
l=[]
m=[]
bb2=[]
scheme1=[]
for a in range(len(bb1)):
    if bb1[a]>10:
        dashu.append(a)
for i in range(len(dashu)-1):
    l.append((bb1[dashu[i]:dashu[i+1]]))
    m.append((scheme[dashu[i]:dashu[i+1]]))
l.append(bb1[dashu[i+1]:])
m.append(scheme[dashu[i+1]:])
# print(l)
# print(m)
for i in range(len(l)-1):
    if l[i]==l[i+1]:
        repeat=i
for s in range(repeat+1):
    bb2=bb2+l[s]
for aaaa in range(repeat):
    scheme1=scheme1+m[aaaa]#无重复
print(bb2)#bb2为order调整后的灯色显示时间
# print(scheme1)
rest=m[repeat]#前重复
add=m[repeat+1]#后重复
changelist = []
addlist = []
for i in range(len(rest)):
    for abc in add[i]:
        changelist.append(abc)
    for location in range(len(rest[i])):
        if rest[i][location] != "r":
            changelist[location] = rest[i][location]
        add1 = "".join(changelist)
    changelist.clear()
    addlist.append(add1)
scheme2 = scheme1 + addlist
print(scheme2)  # scheme2为order调整后的灯色显示顺序
